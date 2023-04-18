import openpyxl

# Excelファイルの読み込み
workbook  = openpyxl.load_workbook('C:/Users/natsume/Documents/python_ocr_excel/sample.xlsx')

### ワークシート名の一覧取得
print(workbook.sheetnames)

### ワークシート名を変数へ格納する
s_name = workbook.sheetnames

sheet = workbook['Sheet1']

### 範囲データ取得
sheet_range = sheet['A1':'A200']


############################### セルの値を取得 ###############################
for row in sheet_range:
    for cell in row:
        # 該当セルの値取得
        cell_value = cell.value

        if cell_value is not None:
            print(cell.coordinate, cell_value)


# Excelファイルを閉じる
workbook.close()