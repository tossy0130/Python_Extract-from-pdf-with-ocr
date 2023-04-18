from io import StringIO
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
import xlsxwriter 

# pip install xlsxwriter

### Excel読み込み
import openpyxl
import pandas as pd

# pip install pandas


# from pdfminer.high_level import extract_text

#実行ファイルのパスを取得
# pwd = os.getcwd()
# print(pwd)


rsrcmgr = PDFResourceManager()
outfp = StringIO()

# laparams の設定
laparams = LAParams()
laparams.detect_vertical = True # Trueにすることで綺麗にテキストを抽出できる

device = TextConverter(rsrcmgr, outfp, codec='utf-8', laparams=laparams)

# pdf_filename = "sample.pdf"
pdf = open('sample.pdf', 'rb')

get_str = ''
get_list = []
interpreter = PDFPageInterpreter(rsrcmgr, device)

############### PDF を　文字列へ出力
for page in PDFPage.get_pages(pdf):
    interpreter.process_page(page)

    str = outfp.getvalue()
    get_str += str

pdf.close()
device.close()
outfp.close()
#print(get_list)

#######　改行コードの削除
a_str = ''
a_str = get_str.replace('\n', '')


####=======####=======
####== 出力
####=======####=======
# print("出力：：：a_str" + a_str)


####### 空白行でリスト作成
a_list = []
a_list = a_str.split(" ")

a_list_r = []

for a in a_list:
    if a != '':
        a_list_r.append(a)

### PDF 読み取り
# text = extract_text(pdf_filename,laparams)

####=======####=======
####== 出力
####=======####=======
"""
for get_pdx in a_list:
    print(get_pdx)
"""

################################ Excel 書き込み ################################
# Excelファイルの読み込み

# wp = openpyxl.Workbook()

#########################  Excel　writer
book = xlsxwriter.Workbook('C:/Users/natsume/Documents/python_ocr_excel/sample.xlsx')
sheet = book.add_worksheet('Sheet1')

# シート名指定
# ws = book.['Sheet1']

# sheet = wp.active

"""
sheet['A1'] = a_list_r[0]
sheet['B1'] = a_list_r[1]
sheet['C1'] = a_list_r[2]
sheet['D1'] = a_list_r[3]
sheet['E1'] = a_list_r[4]
sheet['F1'] = a_list_r[5]

sheet['G1'] = a_list_r[6]
sheet['H1'] = a_list_r[7]
sheet['I1'] = a_list_r[8]
sheet['J1'] = a_list_r[9]
sheet['K1'] = a_list_r[10]
sheet['L1'] = a_list_r[11]

sheet['A2'] = a_list_r[12]
sheet['A3'] = a_list_r[13]
sheet['A4'] = a_list_r[14]
sheet['A5'] = a_list_r[15]
sheet['A6'] = a_list_r[16]
sheet['A7'] = a_list_r[17]

sheet['A8'] = a_list_r[18]
sheet['A9'] = a_list_r[19]
sheet['A10'] = a_list_r[20]
sheet['A11'] = a_list_r[21]
sheet['A12'] = a_list_r[22]
sheet['A13'] = a_list_r[23]
"""

"""
for i in range(len(a_list_r)):
    sheet.cell(row=i + 1, column=1).value = a_list_r[i]
"""

for i in range(len(a_list_r)):
    sheet.write(i + 1,0, a_list_r[i])

book.close()

# セルの A1 
# sheet.cell(row=1, column=1).vlaue = 123

# print("出力：：：：" + a_list_r)

### Excel保存

# wp.save('C:/Users/natsume/Documents/python_ocr_excel/sample.xlsx')

